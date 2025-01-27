import openpyxl, os, peakutils, copy
import numpy as np
import matplotlib.pyplot as plt
from openpyxl.styles.fonts import Font
from openpyxl.styles import PatternFill
from more_itertools import chunked
from scipy import stats

# If the data are not in the same working directory, change the path using the line below.
#os.chdir(r'C:\Users\Hiroshi\Documents\Python_Scripts')

# filename1 must be the text file for green fluorescence. filename2 must be the text file for red fluorescence.
filename1 = 'cfb91d13psi10cf1Gch.txt'
filename2 = 'cfb91d13psi10cf1Rch.txt'

nb = 3 # The Nth block of 10 trials (1 to 6).
nf = 60 # The number of frames in a single trial.
thres_eli = 1.15 # The threshold for eliminating detected peaks that are too small. This number x standard deviation (SD) of 
              # deltaG/R is the threshold. If the height of a peak (from the bottom to the top) is smaller than
              # the threshold, the peak is eliminated.
thres_iso = 1 # The threshold for detecting isolated peaks. This number x SD of deltaG/R is the threshold. 
              # An isolated peak must start from a y-value below the threshold. 
thres_sin = 1.5 # The threshold for detecting peaks that are presumably caused by a single action potential.
               # This number x SD of deltaG/R is the threshold. If the height of a peak is smaller than 
               # the threshold, the peak is considered to be caused by a single action potential.
frame_rate = 61 # The time (in millisecond) to capture a single frame of the movie.
us_onset = 1400 # The time (in millisecond) when the US was delivered.
#cs1 = 1 # The trial number of the first CS only trial.
#us1 = 6 # The trial number of the first US only trial.
#interval = 10 # The interval (i.e., the number of trials) between cs only trials. This is the same as the interval
              # between us only trials. 
              # The code does not work properly if the interval is not a factor of 10.
analysis_complete = 1 # After completion of the analysis, change this value to '1' and run the code again.
                      # This action is necessary to organize the data in Excel Workbook.

# Read two text files then create the list g (green fluorescence intensity) and r (red fluorescence intensity).
with open(filename1) as f1:
    g = []
    for frame1 in f1:
        frame1 = frame1.strip('\n')
        g.append(float(frame1))

with open(filename2) as f2:
    r = []
    for frame2 in f2:
        frame2 = frame2.strip('\n')
        r.append(float(frame2))

# Convert the lists g and r to numpy arrays to detect the baseline with peakutils.
# Obtain the mean value of the baseline for the green fluorescence (mean_b).
g = np.asarray(g)
r = np.asarray(r)
base = peakutils.baseline(g, 10)
mean_b = np.mean(base)

# Smooth r to avoid the potential contribution of green fluorescence bleeding-through to the red channel.
window = 5
edge = int((window-1)/2)
b = np.ones(window)/window
r2 = np.convolve(r, b, mode='valid')
r3 = np.r_[r[:edge], r2, r[-(edge):]]

# Make new lists that contain sublists of g and r. Each sublist contains y-value for a block of 10 trials.
# nb = the Nth block of 10 trials (1 to 6).
# nf = the number of frames in a single trial.
g_chunked = list(chunked(g, nf*10))
g_chunked[nb-1] = np.asarray(g_chunked[nb-1])
base_chunked = peakutils.baseline(g_chunked[nb-1], 5)

r_chunked = list(chunked(r, nf*10))
r_chunked[nb-1] = np.asarray(r_chunked[nb-1])
r3_chunked = list(chunked(r3, nf*10))
r3_chunked[nb-1] = np.asarray(r3_chunked[nb-1])

# Calculate deltaG/R (df). Subtract the baseline value (base_chunked) from g value at each point then divide it by r3.
# Convert the list df to numpy array and detect the baseline of deltaG/R (base_df) and the mean of base_df (mean_bdf).
df = (g_chunked[nb-1]-base_chunked)/r3_chunked[nb-1]
base_df = peakutils.baseline(df, 3)
mean_bdf = np.mean(base_df)

# The standard deviation of deltaG/R.
z = np.std(df)

# Calculate dF/F0 (dfg) and detect the baseline of dF/F0 (base_dfg) for comarison with deltaG/R (df) .
dfg = (g_chunked[nb-1]-base_chunked)/mean_b
base_dfg = peakutils.baseline(dfg, 3)

# Detect peak positions (x-values) in deltaG/R. 
indices = peakutils.indexes(df, thres=0.0025, min_dist=1, thres_abs=True)
    
# Move the peak position if deltaG/R is larger at one frame after a detected peak. 
indices2 = list(indices)
for i in range(0, len(indices)):
    if df[indices[i]+1] > df[indices[i]]:
        indices2.append(indices[i]+1)
        indices2.remove(indices[i])
    elif indices[i] == len(indices) - 1:
        break

indices2.sort()
indices2 = np.asarray(indices2)
   
# Remove peaks that are too small.
indices_all = list(indices2)
threshold1 = z*thres_eli
    
for i in range(0, len(indices2)):
    if indices2[i] <= 1:
        indices_all.remove(indices2[i])
    elif df[indices2[i]-2] > df[indices2[i]-1]:
        if df[indices2[i]] <= df[indices2[i]-1] + threshold1:
            indices_all.remove(indices2[i])
    elif df[indices2[i]-2] <= df[indices2[i]-1]:
        if df[indices2[i]] <= df[indices2[i]-2] + threshold1:
            indices_all.remove(indices2[i])
    
indices_all = np.asarray(indices_all)
    
# Remove peaks that are not well isolated from the preceding peaks.
indices_sp = list(indices_all)
threshold2 = z*thres_iso # How high from the mean baseline (mean_bdf)
    
for i in range(0, len(indices_all)):
    if df[indices_all[i]-2] > df[indices_all[i]-1]:
        if df[indices_all[i]-1] > mean_bdf + threshold2:
            indices_sp.remove(indices_all[i])
    elif df[indices_all[i]-2] <= df[indices_all[i]-1]:
        if df[indices_all[i]-2] > mean_bdf + threshold2:
            indices_sp.remove(indices_all[i])
    
indices_sp = np.asarray(indices_sp)
    
# Calculate the height (from the bottom to the top of the peak) of each isolated peak.
height_sp = []
for i in range(0, len(indices_sp)):
    if df[indices_sp[i]-2] > df[indices_sp[i]-1]:
        height_sp.append(df[indices_sp[i]]-df[indices_sp[i]-1])
    elif df[indices_sp[i]-2] <= df[indices_sp[i]-1]:
        height_sp.append(df[indices_sp[i]]-df[indices_sp[i]-2])
    
height_sp = np.asarray(height_sp)
mean_sp = np.mean(height_sp)
    
# Detect peaks that are presumably caused by a single action potential.
indices_ap1 = list(indices_sp)
threshold3 = z*thres_sin

for i in range(0, len(indices_sp)):
    if df[indices_sp[i]-2] > df[indices_sp[i]-1]:
        if df[indices_sp[i]] > df[indices_sp[i]-1] + threshold3:
            indices_ap1.remove(indices_sp[i])            
    elif df[indices_sp[i]-2] <= df[indices_sp[i]-1]:
        if df[indices_sp[i]] > df[indices_sp[i]-2] + threshold3:
            indices_ap1.remove(indices_sp[i])
    
indices_ap1 = np.asarray(indices_ap1) 

# Define the function to obtain the height of particular types of peaks.
def height(indices_type, height_type):
    for value in indices_type:
        index = indices_sp.index(value)
        height_type.append(height_sp[index])
           
# Obtain the height of isolated peaks that are presumably caused by a single action potential and obtain the average.
height_ap1 = []
indices_sp = list(indices_sp)

height(indices_ap1, height_ap1)
height_ap1 = np.asarray(height_ap1) 
mean_ap1 = np.mean(height_ap1)

# Define the function to sort peaks that occurred during CS period, US period, and all other times.
# CS is delivered at 1200 ms. Time window for CS detection is 1250-1400 ms.
# US is delivered at us_onset (ms). Time window for US detection is (us_onset+25) - (us_onset+150) ms.
# Be careful. The end frame of the detection window depends on whether the end time is divisible by the frame rate.
def sort(list1, list2, list3, list4, list5, list6, list7):
    if 1400 % frame_rate == 0 and (us_onset + 150) % frame_rate == 0:
        for i in range(0, len(list1)):
            for j in range(1, 11):
                if list1[i] >= int(1250/frame_rate) + nf*(j-1) and list1[i] < 1400/frame_rate + nf*(j-1):
                    list2.append(list1[i])
                    list3.append(j + 10*(nb-1))
                elif list1[i] >= int((us_onset + 25)/frame_rate) + nf*(j-1) and list1[i] < (us_onset + 150)/frame_rate + nf*(j-1):
                    list4.append(list1[i])
                    list5.append(j + 10*(nb-1))
                elif ((list1[i] >= 0 + nf*(j-1) and list1[i] < int(1250/frame_rate) + nf*(j-1))
                or (list1[i] >= 1400/frame_rate + nf*(j-1) and list1[i] < int((us_onset + 25)/frame_rate) + nf*(j-1)) 
                or (list1[i] >= (us_onset + 150)/frame_rate + nf*(j-1) and list1[i] <= nf-1 + nf*(j-1))):    
                    list6.append(list1[i])
                    list7.append(j + 10*(nb-1))
        
    elif 1400 % frame_rate == 0 and (us_onset + 150) % frame_rate != 0:
        for i in range(0, len(list1)):
            for j in range(1, 11):
                if list1[i] >= int(1250/frame_rate) + nf*(j-1) and list1[i] < 1400/frame_rate + nf*(j-1):
                    list2.append(list1[i])
                    list3.append(j + 10*(nb-1))
                elif list1[i] >= int((us_onset + 25)/frame_rate) + nf*(j-1) and list1[i] <= (us_onset + 150)/frame_rate + nf*(j-1):
                    list4.append(list1[i])
                    list5.append(j + 10*(nb-1))
                elif ((list1[i] >= 0 + nf*(j-1) and list1[i] < int(1250/frame_rate) + nf*(j-1))
                or (list1[i] >= 1400/frame_rate + nf*(j-1) and list1[i] < int((us_onset + 25)/frame_rate) + nf*(j-1)) 
                or (list1[i] > (us_onset + 150)/frame_rate + nf*(j-1) and list1[i] <= nf-1 + nf*(j-1))):    
                    list6.append(list1[i])
                    list7.append(j + 10*(nb-1))
        
    elif 1400 % frame_rate != 0 and (us_onset + 150) % frame_rate == 0:
        for i in range(0, len(list1)):
            for j in range(1, 11):
                if list1[i] >= int(1250/frame_rate) + nf*(j-1) and list1[i] <= 1400/frame_rate + nf*(j-1):
                    list2.append(list1[i])
                    list3.append(j + 10*(nb-1))
                elif list1[i] >= int((us_onset + 25)/frame_rate) + nf*(j-1) and list1[i] < (us_onset + 150)/frame_rate + nf*(j-1):
                    list4.append(list1[i])
                    list5.append(j + 10*(nb-1))
                elif ((list1[i] >= 0 + nf*(j-1) and list1[i] < int(1250/frame_rate) + nf*(j-1))
                or (list1[i] > 1400/frame_rate + nf*(j-1) and list1[i] < int((us_onset + 25)/frame_rate) + nf*(j-1))  
                or (list1[i] >= (us_onset + 150)/frame_rate + nf*(j-1) and list1[i] <= nf-1 + nf*(j-1))):    
                    list6.append(list1[i])
                    list7.append(j + 10*(nb-1))
                    
    elif 1400 % frame_rate != 0 and (us_onset + 150) % frame_rate != 0:
        for i in range(0, len(list1)):
            for j in range(1, 11):
                if list1[i] >= int(1250/frame_rate) + nf*(j-1) and list1[i] <= 1400/frame_rate + nf*(j-1):
                    list2.append(list1[i])
                    list3.append(j + 10*(nb-1))
                elif list1[i] >= int((us_onset + 25)/frame_rate) + nf*(j-1) and list1[i] <= (us_onset + 150)/frame_rate + nf*(j-1):
                    list4.append(list1[i])
                    list5.append(j + 10*(nb-1))
                elif ((list1[i] >= 0 + nf*(j-1) and list1[i] < int(1250/frame_rate) + nf*(j-1))
                or (list1[i] > 1400/frame_rate + nf*(j-1) and list1[i] < int((us_onset + 25)/frame_rate) + nf*(j-1)) 
                or (list1[i] > (us_onset + 150)/frame_rate + nf*(j-1) and list1[i] <= nf-1 + nf*(j-1))):    
                    list6.append(list1[i])
                    list7.append(j + 10*(nb-1))

    # Lines 211-231 are necessary to reclassify the second peaks in the CS or US time window as spontaneous events.
    list2a = np.asarray(list2)  
    list3a = np.asarray(list3)  
    list4a = np.asarray(list4)   
    list5a = np.asarray(list5)                   
    
    for k in range(1, len(list2a)):
        if list2a[k]-list2a[k-1] <= 150/frame_rate:
            list2.remove(list2a[k])
            list3.remove(list3a[k])
            list6.append(list2a[k])
            list7.append(list3a[k])
            
    for l in range(1, len(list4a)):
        if list4a[l]-list4a[l-1] <= 125/frame_rate:
            list4.remove(list4a[l])
            list5.remove(list5a[l])
            list6.append(list4a[l])
            list7.append(list5a[l])

    list6.sort()
    list7.sort()

indices_all_cs = []
indices_all_cs_trial_number = []
indices_all_us = []
indices_all_us_trial_number = []
indices_all_other = []
indices_all_other_trial_number = []

sort(indices_all, 
     indices_all_cs, indices_all_cs_trial_number, 
     indices_all_us, indices_all_us_trial_number, 
     indices_all_other, indices_all_other_trial_number)   
                
# Detect peaks in CS time window in US only trials and peaks in US time window in CS only trials.
# Reclassify those peaks as other peaks (spontanous peaks).
#def sort2(list8, list9, list10, list11, stim_type):
    #a = np.asarray(list8)
    #for i in range(0, len(a)):
        #for j in range(0, int(10/interval)):
            #if a[i] == stim_type + interval*j + 10*(nb-1):
                #list10.append(list8[i])
                #list11.append(list9[i])
                #list8.remove(list8[i])
                #list9.remove(list9[i])

    #list10.sort()
    #list11.sort()

# Sort all peaks that occurred during CS period, US period, and all other times.
indices_all_cs = []
indices_all_cs_trial_number = []
indices_all_us = []
indices_all_us_trial_number = []
indices_all_other = []
indices_all_other_trial_number = []

sort(indices_all, 
     indices_all_cs, indices_all_cs_trial_number, 
     indices_all_us, indices_all_us_trial_number, 
     indices_all_other, indices_all_other_trial_number)          

#sort2(indices_all_cs_trial_number, indices_all_cs, 
      #indices_all_other_trial_number, indices_all_other, us1)

#sort2(indices_all_us_trial_number, indices_all_us, 
      #indices_all_other_trial_number, indices_all_other, cs1)

# Sort isolated peaks that occurred during CS period, US period, and all other times.
indices_sp_cs = []
indices_sp_cs_trial_number = []
indices_sp_us = []
indices_sp_us_trial_number = []
indices_sp_other = []
indices_sp_other_trial_number = []

sort(indices_sp, 
     indices_sp_cs, indices_sp_cs_trial_number, 
     indices_sp_us, indices_sp_us_trial_number, 
     indices_sp_other, indices_sp_other_trial_number)

#sort2(indices_sp_cs_trial_number, indices_sp_cs, 
      #indices_sp_other_trial_number, indices_sp_other, us1)

#sort2(indices_sp_us_trial_number, indices_sp_us, 
      #indices_sp_other_trial_number, indices_sp_other, cs1)
        
# Define the function to obtain y-values of peaks.
def peak_y(indices_type, y_type):
    for i in indices_type:
       y_type.append(df[i])
    
# Obtain y-values of all detected peaks.
df_all = []
peak_y(indices_all, df_all)

# Obtain y-values of all detected peaks during CS and add 3 for easier visualization.
df_all_cs = []
peak_y(indices_all_cs, df_all_cs)
df_all_cs_add3 = list(map(lambda x: x * 1.1, df_all_cs))

# Obtain y-values of all detected peaks during US and add 3 for easier visualization.
df_all_us = []
peak_y(indices_all_us, df_all_us)
df_all_us_add3 = list(map(lambda x: x * 1.1, df_all_us))
    
# Obtain y-values of isolated peaks.
df_sp = []
peak_y(indices_sp, df_sp)

# Obtain y-values of isolated peaks during CS and add 3 for easier visualization.
df_sp_cs = []
peak_y(indices_sp_cs, df_sp_cs)
df_sp_cs_add3 = list(map(lambda x: x * 1.1, df_sp_cs))

# Obtain y-values of isolated peaks during US and add 3 for easier visualization.
df_sp_us = []
peak_y(indices_sp_us, df_sp_us)
df_sp_us_add3 = list(map(lambda x: x * 1.1, df_sp_us))

# Obtain y-values of isolated peaks that are presumably caused by a single action potential.
df_ap1 = []
peak_y(indices_ap1, df_ap1)

# Obtain the height of isolated peaks that occurred during CS and obtain the average.
height_cs = []
height(indices_sp_cs, height_cs)

height_cs = np.asarray(height_cs) 
mean_cs = np.mean(height_cs)

# Obtain the height of isolated peaks that occurred during US and obtain the average.
height_us = []
height(indices_sp_us, height_us)

height_us = np.asarray(height_us)
mean_us = 0 
if len(height_us) == 0:
    mean_us = None
else:
    mean_us = np.mean(height_us)

# Obtain the height of isolated peaks that occurred all the other times and obtain the average.
height_other = []
height(indices_sp_other, height_other)

height_other = np.asarray(height_other) 
mean_other = np.mean(height_other)

# Plot raw green fluorescence and baseline.
x_low = 0
x_high = nf*10-1 
plt.figure(figsize=(20,4))
plt.ylabel('Green fluorescence')
plt.xlim(x_low, x_high)
plt.plot(g_chunked[nb-1], c='green')
plt.plot(base_chunked, c='orange', label='Baseline')
plt.legend()

# Plot raw and smoothed red fluorescence.
plt.figure(figsize=(20,4))
plt.ylabel('Red fluorescence')
plt.xlim(x_low, x_high)
plt.plot(r_chunked[nb-1], c='red')
plt.plot(r3_chunked[nb-1], c='blue')

# Plot df/F0 and baseline.
plt.figure(figsize=(20,4))
plt.ylabel('dF/F0')
plt.xlim(x_low, x_high)
plt.plot(dfg, c='y')
plt.plot(base_dfg, c='orange', label='Baseline of dF/F0')
plt.legend()

# Plot deltaG/R, baseline, and all detected peaks. Show peaks during CS and US.
plt.figure(figsize=(20,4))
plt.ylabel('deltaG/R')
plt.xlim(x_low, x_high)
plt.plot(df, c='y')
plt.scatter(indices_all, df_all, c = 'green', label = 'All peaks')
plt.scatter(indices_all_cs, df_all_cs_add3, c='silver', marker='v', label='CS')
plt.scatter(indices_all_us, df_all_us_add3, c='sandybrown', marker='v', label='US')
plt.plot(base_df, c='orange', label='Baseline of deltaG/R')
plt.legend()
    
# Plot isolated peaks on deltaG/R. Show peaks during CS and US. 
plt.figure(figsize = (20, 4))
plt.ylabel('deltaG/R')
plt.xlim(x_low, x_high)
plt.plot(df, c='y')
plt.scatter(indices_sp, df_sp, c='red', label='Isolated peaks')
plt.scatter(indices_ap1, df_ap1, c='blue', label='Single AP')
plt.scatter(indices_sp_cs, df_sp_cs_add3, c='silver', marker='v', label='CS')
plt.scatter(indices_sp_us, df_sp_us_add3, c='sandybrown', marker='v', label='US')
plt.legend()
    
plt.show()

# Define the function to write data into the Excel Workbook. 
def write_data():
    sheet1.cell(row=1, column=nb+1).value = nb
    sheet1.cell(row=2, column=nb+1).value = nf
    sheet1.cell(row=3, column=nb+1).value = z
    sheet1.cell(row=4, column=nb+1).value = thres_eli
    sheet1.cell(row=5, column=nb+1).value = threshold1
    sheet1.cell(row=6, column=nb+1).value = thres_iso
    sheet1.cell(row=7, column=nb+1).value = threshold2
    sheet1.cell(row=8, column=nb+1).value = thres_sin  
    sheet1.cell(row=9, column=nb+1).value = threshold3
    sheet1.cell(row=10, column=nb+1).value = frame_rate
    sheet1.cell(row=11, column=nb+1).value = us_onset
    #sheet1.cell(row=12, column=nb+1).value = cs1 
    #sheet1.cell(row=13, column=nb+1).value = us1
    #sheet1.cell(row=14, column=nb+1).value = interval
    sheet1.cell(row=16, column=nb+1).value = len(indices_all)
    sheet1.cell(row=17, column=nb+1).value = len(indices_all)*1000/(nf*10*frame_rate)
    sheet1.cell(row=18, column=nb+1).value = len(indices_all_cs)
    sheet1.cell(row=19, column=nb+1).value = len(indices_all_us)
    sheet1.cell(row=20, column=nb+1).value = len(indices_all_other)
    sheet1.cell(row=22, column=nb+1).value = len(indices_sp)
    sheet1.cell(row=23, column=nb+1).value = len(indices_ap1)
    sheet1.cell(row=24, column=nb+1).value = len(indices_sp_cs)
    sheet1.cell(row=25, column=nb+1).value = len(indices_sp_us)
    sheet1.cell(row=26, column=nb+1).value = len(indices_sp_other)      
    sheet1.cell(row=28, column=nb+1).value = mean_sp
    sheet1.cell(row=29, column=nb+1).value = mean_ap1
    sheet1.cell(row=30, column=nb+1).value = mean_cs
    sheet1.cell(row=31, column=nb+1).value = mean_us
    sheet1.cell(row=32, column=nb+1).value = mean_other 

    for i in range(0, len(df)):
        sheet2.cell(row=i+2, column=nb+4).value = df[i]
        if sheet2.cell(row=i+1, column=1).value in list(indices_all):
            sheet2.cell(row=i+2, column=nb+4).fill = PatternFill(start_color="80f475", fill_type = "solid")
            
    for i in range(0, len(df)):
        if sheet2.cell(row=i+1, column=1).value in indices_all_cs:
            sheet2.cell(row=i+2, column=nb+4).fill = PatternFill(start_color="afaeae", fill_type = "solid")
        
    for i in range(0, len(df)):        
        if sheet2.cell(row=i+1, column=1).value in indices_all_us:
            sheet2.cell(row=i+2, column=nb+4).fill = PatternFill(start_color="ffa477", fill_type = "solid")
    
    for i in range(0, len(df)):
        sheet2.cell(row=i+2, column=nb+12).value = df[i]
        if sheet2.cell(row=i+1, column=1).value in list(indices_sp):
            sheet2.cell(row=i+2, column=nb+12).fill = PatternFill(start_color="fc3f3f", fill_type = "solid")
   
    for i in range(0, len(df)):
        if sheet2.cell(row=i+1, column=1).value in indices_sp_cs:
            sheet2.cell(row=i+2, column=nb+12).fill = PatternFill(start_color="afaeae", fill_type = "solid")
        
    for i in range(0, len(df)):        
        if sheet2.cell(row=i+1, column=1).value in indices_sp_us:
            sheet2.cell(row=i+2, column=nb+12).fill = PatternFill(start_color="ffa477", fill_type = "solid")
   
    for i in range(1, sheet3.max_row + 1):
        for j in range(1, 13):
            sheet3.cell(row=i+1, column=j+13*(nb-1)).value = None
    
    for i in range(0, len(height_cs)):        
        sheet3.cell(row=i+2, column=1+13*(nb-1)).value = indices_sp_cs_trial_number[i]
        sheet3.cell(row=i+2, column=2+13*(nb-1)).value = height_cs[i]
        sheet3.cell(row=i+2, column=3+13*(nb-1)).value = height_cs[i] / mean_ap1
        sheet3.cell(row=i+2, column=4+13*(nb-1)).value = height_cs[i] / mean_other       
    
    for i in range(0, len(height_us)):        
        sheet3.cell(row=i+2, column=5+13*(nb-1)).value = indices_sp_us_trial_number[i]
        sheet3.cell(row=i+2, column=6+13*(nb-1)).value = height_us[i]
        sheet3.cell(row=i+2, column=7+13*(nb-1)).value = height_us[i] / mean_ap1    
        sheet3.cell(row=i+2, column=8+13*(nb-1)).value = height_us[i] / mean_other      
    
    for i in range(0, len(height_other)):        
        sheet3.cell(row=i+2, column=9+13*(nb-1)).value = indices_sp_other_trial_number[i]
        sheet3.cell(row=i+2, column=10+13*(nb-1)).value = height_other[i]
        sheet3.cell(row=i+2, column=11+13*(nb-1)).value = height_other[i] / mean_ap1    
        sheet3.cell(row=i+2, column=12+13*(nb-1)).value = height_other[i] / mean_other     
   
    wb.save(exl_filename)

# Create Excel Workbook (if it is not existed), format, and write data
slice = filename1[0:-4]
exl_filename = slice + '.xlsx'
exl_filename_path = '.\\'
path = exl_filename_path + exl_filename    

if os.path.isfile(path) == False:

    wb = openpyxl.Workbook()
    sheet1 = wb.active
    sheet1.title = 'Summary'
    wb.create_sheet(title='Peaks')
    wb.create_sheet(title='Peak Height')
    wb.create_sheet(title='Peak Height Sorted')
    wb.create_sheet(title='Peak Height Sorted 2')
    wb.create_sheet(title='Results')
    sheet1 = wb['Summary']
    sheet2 = wb['Peaks']
    sheet3 = wb['Peak Height']
    sheet4 = wb['Peak Height Sorted']
    sheet5 = wb['Peak Height Sorted 2']
    sheet6 = wb['Results'] 

    sheet1.column_dimensions['A'].width = 40
    sheet1.cell(row=1, column=1).value = 'Nth block of 10 trials (nb)'
    sheet1.cell(row=2, column=1).value = 'Number of frames per trial (nf)'
    sheet1.cell(row=3, column=1).value = 'SD of deltaG/R'
    sheet1.cell(row=4, column=1).value = 'thres_eli'
    sheet1.cell(row=5, column=1).value = 'Peak elimination threshold'
    sheet1.cell(row=6, column=1).value = 'thres_iso'
    sheet1.cell(row=7, column=1).value = 'Peak isolation threshold'
    sheet1.cell(row=8, column=1).value = 'thres_sin'   
    sheet1.cell(row=9, column=1).value = 'Single AP peak threshold'
    sheet1.cell(row=10, column=1).value = 'Frame rate (ms/frame)'
    sheet1.cell(row=11, column=1).value = 'US onset (ms)' 
    #sheet1.cell(row=12, column=1).value = 'The first CS-only trial'  
    #sheet1.cell(row=13, column=1).value = 'The first US-only trial'  
    #sheet1.cell(row=14, column=1).value = 'The interval between CS (or US) only trials'
    sheet1.cell(row=16, column=1).value = 'Number of all peaks'
    sheet1.cell(row=17, column=1).value = 'Frequency of Ca events (Hz)'
    sheet1.cell(row=18, column=1).value = 'Number of all CS-evoked peaks' 
    sheet1.cell(row=19, column=1).value = 'Number of all US-evoked peaks'
    sheet1.cell(row=20, column=1).value = 'Number of all spontaneous peaks'
    sheet1.cell(row=22, column=1).value = 'Number of isolated peaks'
    sheet1.cell(row=23, column=1).value = 'Number of single AP peaks' 
    sheet1.cell(row=24, column=1).value = 'Number of isolated CS-evoked peaks' 
    sheet1.cell(row=25, column=1).value = 'Number of isolated US-evoked peaks' 
    sheet1.cell(row=26, column=1).value = 'Number of isolated spontaneous peaks'       
    sheet1.cell(row=28, column=1).value = 'Average height of all isolated peaks'
    sheet1.cell(row=29, column=1).value = 'Average height of single AP peaks'
    sheet1.cell(row=30, column=1).value = 'Average height of CS-evoked peaks'
    sheet1.cell(row=31, column=1).value = 'Average height of US-evoked peaks'
    sheet1.cell(row=32, column=1).value = 'Average height of spontaneous peaks'
    sheet1.cell(row=34, column=1).value = 'Number of all peaks in the entire session'
    sheet1.cell(row=35, column=1).value = 'Entire duration of imaging (second)'
    sheet1.cell(row=36, column=1).value = 'Overall frequency of Ca events (Hz)'

    sheet2.cell(row=1, column=1).value = 'All peaks'        
    sheet2.cell(row=1, column=10).value = 'Isolated peaks'  
    sheet2.cell(row=1, column=20).value = '(The values indicate deltaG/R)' 
    sheet2.cell(row=2, column=20).value = '(All peaks are shown in green)'
    sheet2.cell(row=3, column=20).value = '(Isolated peaks are shwon in red)' 
    sheet2.cell(row=4, column=20).value = '(CS-evoked peaks are shwon in gray)' 
    sheet2.cell(row=5, column=20).value = '(US-evoked peaks are shwon in orange)' 
    
    sheet2.cell(row=1, column=1).value = 'Frame'        
    sheet2.cell(row=1, column=2).value = 'Trial'  
    sheet2.cell(row=1, column=3).value = 'Time (s)' 
    sheet2.cell(row=1, column=4).value = 'All peaks'
    sheet2.cell(row=1, column=12).value = 'Isolated peaks' 
   
    sheet3.cell(row=1, column=1).value = '(This sheet is data collection purpose only)'
    sheet3.cell(row=1, column=1).font = Font(color='FFFF0000') 
    
    sheet2.cell(row=1, column=1).font = Font(bold=True)
    sheet2.column_dimensions['A'].width = 6
    sheet2.column_dimensions['B'].width = 6
    sheet2.column_dimensions['C'].width = 8
    sheet2.column_dimensions['D'].width = 10
    sheet2.column_dimensions['L'].width = 13
    
    #for i in range(1, 7):
        #sheet1.cell(row=1, column=i+1).font = Font(bold=True)
    
    for i in range(1, sheet1.max_row + 1):
        sheet1.cell(row=i, column=1).font = Font(bold=True)
    
    for i in range(1, 7):
        sheet2.cell(row=1, column=i+4).value = i
        sheet2.cell(row=1, column=i+12).value = i

    for i in range(1, 19):
        sheet2.cell(row=1, column=i).font = Font(bold=True)
        
    sheet2.cell(row=1, column=4).fill = PatternFill(start_color="80f475", fill_type = "solid")
    sheet2.cell(row=1, column=12).fill = PatternFill(start_color="fc3f3f", fill_type = "solid")       
    
    for i in range(1, len(df)+1):
        sheet2.cell(row=i+1, column=1).value = i
        sheet2.cell(row=i+1, column=1).font = Font(bold=True)
               
    trial_number_per_block = []
    for i in range(1, 11):
         trial_number_per_block.append(i)
    trial_number_per_block = trial_number_per_block * nf
    trial_number_per_block.sort()
    
    for i in range(1, len(df)+1):
            sheet2.cell(row=i+1, column=2).value = trial_number_per_block[i-1]
            sheet2.cell(row=i+1, column=2).font = Font(bold=True)

    time = []
    for i in range(1, nf+1):
         time.append(i*frame_rate/1000)
    time = time * 10
    
    for i in range(1, len(df)+1):
            sheet2.cell(row=i+1, column=3).value = time[i-1]
            sheet2.cell(row=i+1, column=3).font = Font(bold=True)
       
    write_data()

# If the Workbook exists, open it and add data.
else:
    wb = openpyxl.load_workbook(exl_filename)    
    sheet1 = wb['Summary']
    sheet2 = wb['Peaks']
    sheet3 = wb['Peak Height']
    sheet4 = wb['Peak Height Sorted']
    sheet5 = wb['Peak Height Sorted 2']
    sheet6 = wb['Results']     
    write_data()

# In sheet4, organize the data collected in sheet3.
if analysis_complete == 1:
    
    wb.remove(sheet4)
    wb.remove(sheet5)
    wb.remove(sheet6)
    wb.create_sheet(title='Peak Height Sorted')
    wb.create_sheet(title='Peak Height Sorted 2')
    wb.create_sheet(title='Results')
    sheet4 = wb['Peak Height Sorted']
    sheet5 = wb['Peak Height Sorted 2']
    sheet6 = wb['Results']
   
    for i in range(34, 37):
        sheet1.cell(row=i, column=2).value = None
    
    number_all_peaks = []
    for i in range(2, sheet1.max_column + 1):
        number_all_peaks.append(sheet1.cell(row=16, column=i).value) 

    sum_all_peaks = np.sum(np.asarray(number_all_peaks))
    duration = nf*10*(sheet1.max_column-1)*frame_rate/1000

    sheet1.cell(row=34, column=2).value = sum_all_peaks
    sheet1.cell(row=35, column=2).value = duration
    sheet1.cell(row=36, column=2).value = sum_all_peaks/duration  
          
    # Count the number of each type of peaks in each block of 10 trials.
    cs_peak_number = []
    us_peak_number = [] 
    other_peak_number =[]

    for i in range(2, sheet1.max_column + 1):
        cs_peak_number.append(sheet1.cell(row=24, column=i).value)
        us_peak_number.append(sheet1.cell(row=25, column=i).value)
        other_peak_number.append(sheet1.cell(row=26, column=i).value)

    # Count the cumulative number of each type of peaks in each block of 10 trials.
    cs_peak_number2 = copy.copy(cs_peak_number)
    us_peak_number2 = copy.copy(us_peak_number)
    other_peak_number2 =copy.copy(other_peak_number)

    cs_peak_number_sum = []
    us_peak_number_sum = []
    other_peak_number_sum =[]

    for i in range(0, sheet1.max_column-2):
        cs_peak_number_sum.append(cs_peak_number2[i])
        us_peak_number_sum.append(us_peak_number2[i])
        other_peak_number_sum.append(other_peak_number2[i])
      
        cs_peak_number2[i+1] = cs_peak_number2[i] + cs_peak_number2[i+1]
        us_peak_number2[i+1] = us_peak_number2[i] + us_peak_number2[i+1]
        other_peak_number2[i+1] = other_peak_number2[i] + other_peak_number2[i+1]

    # Sort peaks. Arrange CS, US, and other peaks in each column.
    def sort_data1(peak_number_list, start_column):
        for i in range(1, peak_number_list[0]+1):
            for j in range(start_column, start_column+4):
                sheet4.cell(row=i+1, column=j).value = sheet3.cell(row=i+1, column=j).value

    sort_data1(cs_peak_number, 1)
    sort_data1(us_peak_number, 5)
    sort_data1(other_peak_number, 9)
    
    def sort_data2(peak_number_list, peak_number_sum_list, start_column):
        for i in range(1, sheet1.max_column-1):
            for j in range(1, peak_number_list[i]+1):
                for k in range(start_column, start_column+4):
                    sheet4.cell(row=j+1+peak_number_sum_list[i-1], column=k).value = sheet3.cell(row=j+1, column=i*13+k).value
                
    sort_data2(cs_peak_number, cs_peak_number_sum, 1)
    sort_data2(us_peak_number, us_peak_number_sum, 5)
    sort_data2(other_peak_number, other_peak_number_sum, 9)
   
    for i in range(1, 4):
        sheet4.cell(row=1, column=1+4*(i-1)).value = 'Trial #'
       
    sheet4.cell(row=1, column=2).value = 'CS (abs)'
    sheet4.cell(row=1, column=3).value = 'CS (norm to 1AP)'
    sheet4.cell(row=1, column=4).value = 'CS (norm to spon)'    
    sheet4.cell(row=1, column=6).value = 'US (abs)'    
    sheet4.cell(row=1, column=7).value = 'US (norm to 1AP)'
    sheet4.cell(row=1, column=8).value = 'US (norm to spon)'
    sheet4.cell(row=1, column=10).value = 'Spon (abs)'
    sheet4.cell(row=1, column=11).value = 'Spon (norm to 1AP)'
    sheet4.cell(row=1, column=12).value = 'Spon (norm to spon)'   

    for i in range(1, 13):
        sheet4.cell(row=1, column=i).font = Font(bold=True)
    
    sheet4.column_dimensions['A'].width = 6
    sheet4.column_dimensions['B'].width = 18
    sheet4.column_dimensions['C'].width = 18
    sheet4.column_dimensions['D'].width = 18
    sheet4.column_dimensions['E'].width = 6
    sheet4.column_dimensions['F'].width = 18
    sheet4.column_dimensions['G'].width = 18
    sheet4.column_dimensions['H'].width = 18
    sheet4.column_dimensions['I'].width = 6
    sheet4.column_dimensions['J'].width = 18
    sheet4.column_dimensions['K'].width = 18
    sheet4.column_dimensions['L'].width = 18
    
    # Distinguish CS only trials and US only trials from other trials (i.e., both CS and US were presented).
    # First, collect data in each column of sheet4.
    cs_trial_number = []
    cs_norm_1ap = []
    cs_norm_spon = []
    us_trial_number = []
    us_norm_1ap = []
    us_norm_spon = []
    other_trial_number = []
    other_norm_1ap = []
    other_norm_spon = []
    
    for i in range(2, np.sum(np.asarray(cs_peak_number))+2):
        cs_trial_number.append(sheet4.cell(row=i, column=1).value)
        cs_norm_1ap.append(sheet4.cell(row=i, column=3).value)
        cs_norm_spon.append(sheet4.cell(row=i, column=4).value)
        
    for i in range(2, np.sum(np.asarray(us_peak_number))+2):
        us_trial_number.append(sheet4.cell(row=i, column=5).value)
        us_norm_1ap.append(sheet4.cell(row=i, column=7).value) 
        us_norm_spon.append(sheet4.cell(row=i, column=8).value) 
        
    for i in range(2, np.sum(np.asarray(other_peak_number))+2):
        other_trial_number.append(sheet4.cell(row=i, column=9).value)
        other_norm_1ap.append(sheet4.cell(row=i, column=11).value) 
        other_norm_spon.append(sheet4.cell(row=i, column=12).value) 
    
    # Convert the lists to numpy arrays. 
    cs_trial_number = np.asarray(cs_trial_number)
    cs_norm_1ap = np.asarray(cs_norm_1ap)
    cs_norm_spon = np.asarray(cs_norm_spon)
    us_trial_number = np.asarray(us_trial_number)
    us_norm_1ap = np.asarray(us_norm_1ap)
    us_norm_spon = np.asarray(us_norm_spon) 
    
    # Separete CS only trials and US only trials from other trials     
    #cs_alone_trial_number = []
    #cs_alone_norm_1ap = []
    #cs_alone_norm_spon = []
    #cs_with_us_trial_number = list(cs_trial_number)
    #cs_with_us_norm_1ap = list(cs_norm_1ap)   
    #cs_with_us_norm_spon = list(cs_norm_spon)      
    
    #for i in range(0, len(cs_trial_number)):
        #for j in range(0, 6):
            #if cs_trial_number[i] == cs1 + interval * j:
                #cs_alone_trial_number.append(cs_trial_number[i])
                #cs_alone_norm_1ap.append(cs_norm_1ap[i])
                #cs_alone_norm_spon.append(cs_norm_spon[i])
                #cs_with_us_trial_number.remove(cs_trial_number[i])
                #cs_with_us_norm_1ap.remove(cs_norm_1ap[i])                
                #cs_with_us_norm_spon.remove(cs_norm_spon[i])  
    
    #us_alone_trial_number = []
    #us_alone_norm_1ap = []
    #us_alone_norm_spon = []
    #us_with_cs_trial_number = list(us_trial_number)
    #us_with_cs_norm_1ap = list(us_norm_1ap)   
    #us_with_cs_norm_spon = list(us_norm_spon)       
    
    #for i in range(0, len(us_trial_number)):
        #for j in range(0, 6):
            #if us_trial_number[i] == us1 + interval * j:
                #us_alone_trial_number.append(us_trial_number[i])
                #us_alone_norm_1ap.append(us_norm_1ap[i])
                #us_alone_norm_spon.append(us_norm_spon[i])                
                #us_with_cs_trial_number.remove(us_trial_number[i])
                #us_with_cs_norm_1ap.remove(us_norm_1ap[i])
                #us_with_cs_norm_spon.remove(us_norm_spon[i])    
    
    # Write the data in sheet5.
    #def sort_data3(list_type, column_number):
        #for i in range(1, len(list_type)+1):
                #sheet5.cell(row=i+2, column=column_number).value = list_type[i-1]
          
    #sort_data3(cs_alone_trial_number, 1)   
    #sort_data3(cs_alone_norm_1ap, 2) 
    #sort_data3(cs_alone_norm_spon, 3) 
    #sort_data3(cs_with_us_trial_number, 4)    
    #sort_data3(cs_with_us_norm_1ap, 5) 
    #sort_data3(cs_with_us_norm_spon, 6) 
    #sort_data3(us_alone_trial_number, 7)    
    #sort_data3(us_alone_norm_1ap, 8) 
    #sort_data3(us_alone_norm_spon, 9)  
    #sort_data3(us_with_cs_trial_number, 10)      
    #sort_data3(us_with_cs_norm_1ap, 11) 
    #sort_data3(us_with_cs_norm_spon, 12) 
    #sort_data3(other_trial_number, 13) 
    #sort_data3(other_norm_1ap, 14)
    #sort_data3(other_norm_spon, 15)

    #sheet5.cell(row=1, column=1).value = 'CS alone'   
    #sheet5.cell(row=1, column=4).value = 'CS in CS-US trial'    
    #sheet5.cell(row=1, column=7).value = 'US alone'   
    #sheet5.cell(row=1, column=10).value = 'US in CS-US trial'    
    #sheet5.cell(row=1, column=13).value = 'Spontaneous'

    #for i in range(1, 6):
        #sheet5.cell(row=2, column=1+(i-1)*3).value = 'Trial #' 
        #sheet5.cell(row=2, column=2+(i-1)*3).value = 'Norm to 1AP'
        #sheet5.cell(row=2, column=3+(i-1)*3).value = 'Norm to spon'
    
    #for i in range(1, 6):
        #sheet5.cell(row=1, column=1+3*(i-1)).font = Font(bold=True)
    
    #sheet5.column_dimensions['A'].width = 15
    #sheet5.column_dimensions['B'].width = 12
    #sheet5.column_dimensions['C'].width = 12
    #sheet5.column_dimensions['D'].width = 15
    #sheet5.column_dimensions['E'].width = 12
    #sheet5.column_dimensions['F'].width = 12
    #sheet5.column_dimensions['G'].width = 15
    #sheet5.column_dimensions['H'].width = 12
    #sheet5.column_dimensions['I'].width = 12
    #sheet5.column_dimensions['J'].width = 15
    #sheet5.column_dimensions['K'].width = 12
    #sheet5.column_dimensions['L'].width = 12
    #sheet5.column_dimensions['M'].width = 15
    #sheet5.column_dimensions['N'].width = 12
    #sheet5.column_dimensions['O'].width = 12
    
    # Summarize the results of quantification in sheet6.
    sheet6.cell(row=1, column=2).value = 'CS peaks in all trials'   
    sheet6.cell(row=1, column=3).value = 'CS peaks in CS alone trials'
    sheet6.cell(row=1, column=4).value = 'CS peaks in CS-US trials'
    sheet6.cell(row=1, column=5).value = 'US peaks in all trials'    
    sheet6.cell(row=1, column=6).value = 'US peaks in US alone trials'
    sheet6.cell(row=1, column=7).value = 'US peaks in CS-US trials' 
    sheet6.cell(row=1, column=8).value = 'Spontaneous peaks'   
    
    sheet6.cell(row=2, column=1).value = 'Number of isolated events'  
    sheet6.cell(row=4, column=1).value = 'Norm peak to 1AP: height_mean'   
    sheet6.cell(row=5, column=1).value = 'Norm peak to 1AP: height_SEM' 
    sheet6.cell(row=6, column=1).value = 'Norm peak to 1AP: height_SD'  
    sheet6.cell(row=8, column=1).value = 'Norm peak to spon: height_mean'  
    sheet6.cell(row=9, column=1).value = 'Norm peak to spon: height_SEM'   
    sheet6.cell(row=10, column=1).value = 'Norm peak to spon: height_SD'       
    
    #cs_alone_norm_1ap = np.asarray(cs_alone_norm_1ap)
    #cs_with_us_norm_1ap = np.asarray(cs_with_us_norm_1ap)
    #us_alone_norm_1ap = np.asarray(us_alone_norm_1ap)
    #us_with_cs_norm_1ap = np.asarray(us_with_cs_norm_1ap)
    other_norm_1ap = np.asarray(other_norm_1ap)

    #cs_alone_norm_spon = np.asarray(cs_alone_norm_spon)
    #cs_with_us_norm_spon = np.asarray(cs_with_us_norm_spon)
    #us_alone_norm_spon = np.asarray(us_alone_norm_spon)
    #us_with_cs_norm_spon = np.asarray(us_with_cs_norm_spon)
    other_norm_spon = np.asarray(other_norm_spon)

    cs_all_trial_data = [len(cs_norm_1ap), np.mean(cs_norm_1ap), stats.sem(cs_norm_1ap), np.std(cs_norm_1ap, ddof=1),
                         np.mean(cs_norm_spon), stats.sem(cs_norm_spon), np.std(cs_norm_spon, ddof=1)]
    #cs_alone_trial_data = [len(cs_alone_norm_1ap), np.mean(cs_alone_norm_1ap), stats.sem(cs_alone_norm_1ap), np.std(cs_alone_norm_1ap, ddof=1),
                         #np.mean(cs_alone_norm_spon), stats.sem(cs_alone_norm_spon), np.std(cs_alone_norm_spon, ddof=1)]
    #cs_csus_trial_data = [len(cs_with_us_norm_1ap), np.mean(cs_with_us_norm_1ap), stats.sem(cs_with_us_norm_1ap), np.std(cs_with_us_norm_1ap, ddof=1),
                         #np.mean(cs_with_us_norm_spon), stats.sem(cs_with_us_norm_spon), np.std(cs_with_us_norm_spon, ddof=1)]
    us_all_trial_data = [len(us_norm_1ap), np.mean(us_norm_1ap), stats.sem(us_norm_1ap), np.std(us_norm_1ap, ddof=1),
                         np.mean(us_norm_spon), stats.sem(us_norm_spon), np.std(us_norm_spon, ddof=1)]
    #us_alone_trial_data = [len(us_alone_norm_1ap), np.mean(us_alone_norm_1ap), stats.sem(us_alone_norm_1ap), np.std(us_alone_norm_1ap, ddof=1),
                         #np.mean(us_alone_norm_spon), stats.sem(us_alone_norm_spon), np.std(us_alone_norm_spon, ddof=1)]
    #us_csus_trial_data = [len(us_with_cs_norm_1ap), np.mean(us_with_cs_norm_1ap), stats.sem(us_with_cs_norm_1ap), np.std(us_with_cs_norm_1ap, ddof=1),
                         #np.mean(us_with_cs_norm_spon), stats.sem(us_with_cs_norm_spon), np.std(us_with_cs_norm_spon, ddof=1)]
    other_data = [len(other_norm_1ap), np.mean(other_norm_1ap), stats.sem(other_norm_1ap), np.std(other_norm_1ap, ddof=1),
                  np.mean(other_norm_spon), stats.sem(other_norm_spon), np.std(other_norm_spon, ddof=1)] 
    
    def sort_date4(list_type2, column_number2):
        sheet6.cell(row=2, column=column_number2).value = list_type2[0]
        
        for i in range(1, 4):
            sheet6.cell(row=i+3, column=column_number2).value = list_type2[i]

        for i in range(4, 7):
            sheet6.cell(row=i+4, column=column_number2).value = list_type2[i]
    
    sort_date4(cs_all_trial_data, 2)
    #sort_date4(cs_alone_trial_data, 3)            
    #sort_date4(cs_csus_trial_data, 4)
    sort_date4(us_all_trial_data, 5)
    #sort_date4(us_alone_trial_data, 6)
    #sort_date4(us_csus_trial_data, 7)
    sort_date4(other_data, 8)

    for i in range(1, 9):
        sheet6.cell(row=1, column=i).font = Font(bold=True)
    
    for i in range(2, 12):
        sheet6.cell(row=i, column=1).font = Font(bold=True)
        
    sheet6.column_dimensions['A'].width = 29
    sheet6.column_dimensions['B'].width = 18
    sheet6.column_dimensions['C'].width = 23
    sheet6.column_dimensions['D'].width = 23
    sheet6.column_dimensions['E'].width = 18
    sheet6.column_dimensions['F'].width = 24
    sheet6.column_dimensions['G'].width = 22
    sheet6.column_dimensions['H'].width = 18
    
wb.save(exl_filename)