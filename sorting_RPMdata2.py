# You need to download openpyxl module. To do so, from Anaconda ptompt window, conda install -c conda-forge openpyxl
# You also need to set a pass in os.chdir
import openpyxl, os, numpy
from openpyxl.styles.fonts import Font
os.chdir(r'C:\Users\NishiyamaLab\Data')

#Read a text file (Data.txt) and put the data into an Excel file and create two new sheets in the Excel file. 
data = []
with open("rpm1d9w.txt") as f:
    for line in f:
        line = line.strip('\n')
        data.append(float(line))

wb = openpyxl.Workbook()
sheet1 = wb.active
sheet1.title = 'Data'
wb.create_sheet(title = 'Sorted')
wb.create_sheet(title = 'Analyzed')
sheet1 = wb['Data']
sheet2 = wb['Sorted']
sheet3 = wb['Analyzed']

for row_index in range(0, len(data)):
    sheet1.cell(row = row_index + 1, column = 1).value = data[row_index]

# Set the format of sheet3.
sheet3.column_dimensions['A'].width = 12
sheet3.cell(row = 1, column = 1).value = 'Trial #'
sheet3.cell(row = 1, column = 2).value = 'CS mean'
sheet3.cell(row = 1, column = 3).value = 'CS sum'
sheet3.cell(row = 1, column = 4).value = 'US mean'
sheet3.cell(row = 1, column = 5).value = 'US sum'
sheet3.cell(row = 1, column = 6).value = 'Trial mean'
sheet3.cell(row = 1, column = 7).value = 'Trial sum'
sheet3.cell(row = 62, column = 1).value = 'Across Trials'
sheet3.cell(row = 62, column = 1).font = Font(bold = True)

for trial in range (1, 61):
    sheet3.cell(row = trial + 1, column = 1).value = trial
    sheet3.cell(row = trial + 1, column = 1).font = Font(bold = True)

for col in range(1, 8):
    sheet3.cell(row = 1, column = col).font = Font(bold = True)

# Sort the original, 1-dimensionally arranged data in sheet1 to a 2-dimensional array in sheet2. 
# 1500 indicates that TTL is given.
column_counter = 1
row_counter = 1

for i in range(1, sheet1.max_row + 1):
    if sheet1.cell(row=i, column=1).value == 1500:
        column_counter += 1
        sheet2.cell(row=1, column=column_counter).value = sheet1.cell(row=i, column=1).value
        row_counter = 1
    else:
        row_counter += 1
        sheet2.cell(row_counter, column=column_counter).value = sheet1.cell(row=i, column=1).value

# Analyze the data in sheet2 and show the results in sheet3. CS, US, and entire trial (tri) ranges are randomely set for now.
for trial_columns in range(2, sheet2.max_column + 1):
    cs = []
    us = []
    tri = []
    for cs_rows in range(24, 28):
        cs.append(sheet2.cell(row=cs_rows, column=trial_columns).value) 
        sheet3.cell(row=trial_columns, column=2).value = numpy.mean(cs)
        sheet3.cell(row=trial_columns, column=3).value = numpy.sum(cs)
    for us_rows in range(29, 33):
        us.append(sheet2.cell(row=us_rows, column=trial_columns).value)
        sheet3.cell(row=trial_columns, column=4).value = numpy.mean(us)
        sheet3.cell(row=trial_columns, column=5).value = numpy.sum(us)
    for tri_rows in range(2, 251):
        tri.append(sheet2.cell(row=tri_rows, column=trial_columns).value)
        sheet3.cell(row=trial_columns, column=6).value = numpy.mean(tri)
        sheet3.cell(row=trial_columns, column=7).value = numpy.sum(tri)

# Obtain averaged values across trials from results in sheet3.
for data_type in range(2, 8):
    across = []
    for trial_rows in range(2, sheet2.max_column+1):
        across.append(sheet3.cell(row=trial_rows, column=data_type).value)
        sheet3.cell(row=63, column=data_type).value = numpy.mean(across)
        sheet3.cell(row=64, column=data_type).value = numpy.sum(across)

wb.save('rpm1d9w.xlsx')